#Importamos la librería pandas
import pandas as pd

#Importamos la librería NumPy
import numpy as nmp

#Importamos la librería con la función de interactuar con los archivos xlsx
import openpyxl

#Abrimos el archivo que cargamos previamente y activamos la hoja
arch1 = openpyxl.load_workbook('Pandaa1.xlsx', data_only=True)
hoja = arch1.active

#Ubicamos estrictamente la información que queremos utilizar
datos = hoja['A1':'C3']
print("Los datos con su ubicación especificada son: \n", datos, "\n")

#Creamos una lista vacía
matriz1 = []
print("\nLa lista vacía es: \n", matriz1, "\n")

#Utilizamos un for para definir como valores cada fila de datos y agregarlos a la lista
for fila in datos:
    x = [celda.value for celda in fila]
    matriz1.append(x)
print("\nLa matriz 1 es: \n", matriz1, "\n")

#Cerramos el archivo
arch1.close()

#Creamos el archivo para guardar los resultados y activamos la hoja
arch3 = openpyxl.Workbook()
hojitita = arch3.active

#Convertimos la lista en un Dataframe
dfm1 = pd.DataFrame(matriz1)
print("\nLa lista convertida en Dataframe es: \n", dfm1, "\n")

#Añadimos los nombres de las filas y columnas
dfm2 = pd.DataFrame(matriz1, index=[1, 2, 3], columns=["X", "Y", "Z"])
print("\nEl Dataframe con sus filas y columnas: \n", str(dfm2), "\n")

#Cambiamos los datos de una posición en específico
dfm2.iloc[1, 2] = 7
dfm2.iloc[2, 1] = 1
print("\nEl nuevo contenido del Dataframe: \n", (str(dfm2)), "\n")

#Añadimos los datos al documento nuevo
hojitita['B2'] = ("Contenido del archivo")
hojitita['C2'] = (str(matriz1))
hojitita['B3'] = ("Dataframe con contenido del archivo")
hojitita['C3'] = (str(dfm1))
hojitita['B4'] = ("Dataframe con contenido del archivo")
hojitita['C4'] = (str(dfm2))

#Guardamos el archivo nuevo
arch3.save('res1.xlsx')

#Abrimos el archivo que cargamos previamente y activamos la hoja
arch2 = openpyxl.load_workbook('Pandaa2.xlsx', data_only=True)
hojita = arch2.active

#Ubicamos estrictamente la información que queremos utilizar
datitos = hojita['A1':'C3']
print("Los datos con su ubicación especificada son: \n", datitos, "\n")

#Creamos una lista vacía
matriz2 = []
print("\nLa lista vacía es: \n", matriz2, "\n")

#Utilizamos un for para definir como valores cada fila de datos y agregarlos a la lista
for fila in datitos:
    y = [celda.value for celda in fila]
    matriz2.append(y)
print("\nLa matriz 2 es: \n", matriz2, "\n")

#Cerramos el archivo
arch2.close()

#Creamos el archivo para guardar los resultados y activamos la hoja
arch4 = openpyxl.Workbook()
hojititita = arch4.active

#Convertimos la lista en Dataframe
dfm21 = pd.DataFrame(matriz2)
print("\nLa lista convertida en Dataframe es: \n", dfm21, "\n")

#Añadimos los nombres de las filas y columnas
dfm22 = pd.DataFrame(matriz2, index=[1, 2, 3], columns=["F", "G", "H"])
print("\nEl Dataframe con sus filas y columnas: \n", str(dfm22), "\n")

#Creamos una matriz nueva
array1 = nmp.array([[7, 14, 21], [5, 10, 15], [3, 6, 9]])
print("\nLa nueva matriz es: \n", array1, "\n")

#Creamos un Dataframe con la nueva matriz
dfmm1 = pd.DataFrame(array1, index=[7, 8, 9], columns=["U", "I", "O"])
print("\nEl Dataframe de la nueva matriz: \n", dfmm1, "\n")

#Concatenamos dos Dataframes
conc1 = pd.concat([dfm2, dfmm1])
print("\nLa concatenación de dos matrices: \n", conc1, "\n")

#Añadimos los datos al documento nuevo
hojita['B1'] = ("Contenido del archivo")
hojita['C1'] = (str(matriz2))
hojita['B2'] = ("Dataframe con contenido del archivo")
hojita['C2'] = (str(dfm21))
hojita['B3'] = ("Dataframe con contenido de archivo y nombres de celdas")
hojita['C3'] = (str(dfm22))
hojita['B4'] = ("Dataframe con nueva entrada y nombres de celdas")
hojita['C4'] = (str(dfmm1))
hojita['B5'] = ("Concatenacion primer y segundo Dataframe")
hojita['C5'] = (str(conc1))

#Guardamos el archivo
arch4.save('res2.xlsx')
